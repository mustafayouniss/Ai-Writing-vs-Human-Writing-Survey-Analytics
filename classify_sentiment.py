import openpyxl
import re
from collections import Counter

def normalize(text):
    return str(text or '').strip()

def word_count(text):
    return len(text.split()) if text else 0

def count_keywords(text, keywords):
    text_lower = text.lower()
    return sum(1 for kw in keywords if kw in text_lower)

def has_pattern(text, patterns):
    text_lower = text.lower()
    return any(re.search(p, text_lower) for p in patterns)

# ============ KEYWORD LISTS ============

POSITIVE_KW = [
    # Arabic
    'سريع', 'سرعة', 'سرعه', 'توفير', 'مفيد', 'مفيده', 'مفيدة',
    'سهل', 'أسهل', 'اسهل', 'منظم', 'منظمة', 'منظمه',
    'دقيق', 'دقة', 'صحيح', 'صحة', 'قوي', 'جيد', 'ممتاز', 'رائع', 'كويس',
    'كفاءة', 'فعال', 'بيساعد', 'يساعد', 'احسن', 'مريح',
    'جميل', 'عظيم', 'الأفضل', 'حلو', 'ممتازة', 'رائعة',
    # English
    'fast', 'quick', 'speed', 'efficient', 'accurate', 'useful',
    'helpful', 'good', 'great', 'save', 'saving', 'organized',
    'easy', 'convenient', 'reliable', 'productive', 'professional',
    'error-free', 'correct', 'grammar', 'advantage', 'benefit',
    'excellent', 'amazing', 'perfect', 'wonderful', 'best',
    'powerful', 'smart', 'structured', 'time'
]

NEGATIVE_KW = [
    # Arabic
    'تكرار', 'بيكرر', 'تكراري', 'مكرر', 'مكرره',
    'ممل', 'جمود', 'جامد', 'جامدة',
    'خطأ', 'غلط', 'اخطاء', 'أخطاء', 'خطا',
    'عيب', 'عيوب', 'ضعف', 'ضعيف', 'سيء',
    'مشكل', 'مشكلة', 'نقص', 'سطحي', 'سطحية',
    'سرقة', 'نسخ', 'خطير', 'كارثة', 'فظيع',
    'بارد', 'آلي', 'ركيك', 'ركيكة',
    # English
    'repetitive', 'boring', 'cold', 'robotic', 'generic', 'problem',
    'issue', 'error', 'wrong', 'bad', 'limited', 'limitation',
    'mechanical', 'rigid', 'plagiarism', 'terrible', 'awful',
    'worst', 'useless', 'stiff', 'lack', 'lacks', 'lacking'
]

# Emotional keywords (indicate the person has feelings/passion about the topic)
EMOTIONAL_POSITIVE_KW = [
    'ممتاز', 'رائع', 'عظيم', 'جميل', 'حلو', 'amazing', 'love',
    'excellent', 'wonderful', 'fantastic', 'best', 'perfect',
    'جدا', 'كتير', 'very', 'really', 'extremely', 'absolutely',
    'بجد', 'فعلا', 'اوي', 'قوي'
]

EMOTIONAL_NEGATIVE_KW = [
    'روح', 'مشاعر', 'عاطف', 'عاطفة', 'عاطفه', 'احساس', 'إحساس',
    'ابداع', 'إبداع', 'خيال', 'بشري', 'انساني', 'إنساني',
    'emotion', 'feeling', 'soul', 'creativity', 'human touch',
    'originality', 'authentic', 'heart', 'passion',
    'بتدمر', 'يدمر', 'خطير', 'كارثة', 'فظيع', 'سيء جدا',
    'terrible', 'awful', 'horrible', 'destroy', 'dangerous'
]

HEDGING_KW = [
    'ممكن', 'يمكن', 'الى حد ما', 'إلى حد ما', 'نوعا ما',
    'احيانا', 'أحيانا', 'ساعات', 'شوية', 'شويه', 'على حسب',
    'بعض', 'حسب',
    'maybe', 'perhaps', 'sometimes', 'could be', 'sort of',
    'kind of', 'partially', 'to some extent', 'depends',
    'it depends', 'not always', 'not necessarily'
]

# Q2 dismissal patterns (person says AI has NO disadvantages)
DISMISSAL_Q2 = [
    r'^لا\.?\s*$', r'^no\.?\s*$', r'^مفيش', r'^لا مفيش',
    r'ملهاش عيوب', r'لا ملهاش', r'^not really',
    r'^none\s*$', r'^nothing\s*$', r'ليس لها عيوب',
    r'لا اعتقد', r'لا ما اعتقد', r'^مش شايف',
    r'لا ملهاش عيوب', r'^ما اظن'
]

# ============ CLASSIFICATION ============

def classify_row(q1, q2, q3):
    q1 = normalize(q1)
    q2 = normalize(q2)
    q3 = normalize(q3)

    combined = f"{q1} {q2} {q3}".lower()

    # ===== SCORING =====
    pos_score = 0
    neg_score = 0

    # --- Q2: Does the person dismiss AI disadvantages? ---
    dismisses_q2 = has_pattern(q2, DISMISSAL_Q2)
    if dismisses_q2:
        pos_score += 3

    # --- Q3: Advantages (positive context) ---
    q3_lower = q3.lower()
    q3_pos = count_keywords(q3_lower, POSITIVE_KW)
    pos_score += q3_pos
    q3_words = word_count(q3)
    if q3_words >= 15:
        pos_score += 2
    elif q3_words >= 8:
        pos_score += 1

    # --- Q1 & Q2: Criticism intensity ---
    q1q2 = f"{q1} {q2}".lower()
    if not dismisses_q2:
        neg_score += count_keywords(q1q2, NEGATIVE_KW)
        q1q2_words = word_count(q1) + word_count(q2)
        if q1q2_words >= 25:
            neg_score += 2
        elif q1q2_words >= 12:
            neg_score += 1

    # Q1 always has some criticism (question asks what AI lacks)
    neg_score += min(count_keywords(q1.lower(), NEGATIVE_KW), 2)

    # --- Emotional intensity ---
    emo_pos = count_keywords(combined, EMOTIONAL_POSITIVE_KW)
    emo_neg = count_keywords(combined, EMOTIONAL_NEGATIVE_KW)

    # --- Hedging ---
    hedging = count_keywords(combined, HEDGING_KW)

    # ===== GENERAL SENTIMENT =====
    diff = pos_score - neg_score
    if diff >= 2:
        general = 'Positive'
    elif diff <= -2:
        general = 'Negative'
    else:
        general = 'Neutral'

    # ===== DETAILED EMOTION =====

    # Check for Pragmatic: factual, no emotional language, just lists facts
    has_emotion = (emo_pos + emo_neg) >= 2
    is_hedging = hedging >= 2

    if not has_emotion and not is_hedging and hedging <= 1:
        # Pragmatic: no strong emotional words, just stating facts
        if general == 'Positive':
            # Check if they're enthusiastic (Supportive) or just factual (Pragmatic)
            if emo_pos >= 1:
                detailed = 'Supportive'
            else:
                detailed = 'Pragmatic'
        elif general == 'Negative':
            # Check if they're passionate (Critical) or just listing problems (Pragmatic)
            if emo_neg >= 1:
                detailed = 'Critical'
            else:
                detailed = 'Pragmatic'
        else:
            detailed = 'Pragmatic'
    elif is_hedging:
        # Ambivalent: using hedging language, can't decide
        detailed = 'Ambivalent'
    elif general == 'Positive':
        if emo_pos >= 1:
            detailed = 'Supportive'
        elif hedging >= 1:
            detailed = 'Ambivalent'
        else:
            detailed = 'Supportive'
    elif general == 'Negative':
        if emo_neg >= 2 or neg_score >= 5:
            detailed = 'Critical'
        elif hedging >= 1:
            detailed = 'Ambivalent'
        else:
            detailed = 'Concerned'
    elif general == 'Neutral':
        if hedging >= 1:
            detailed = 'Ambivalent'
        elif emo_neg > emo_pos:
            detailed = 'Concerned'
        elif emo_pos > emo_neg:
            detailed = 'Supportive'
        else:
            detailed = 'Ambivalent'

    return general, detailed

# ============ MAIN ============

wb = openpyxl.load_workbook('data/Written Questions.xlsx')
ws = wb.active

# Clear old columns if they exist
for row_idx in range(1, ws.max_row + 1):
    ws.cell(row=row_idx, column=5, value=None)
    ws.cell(row=row_idx, column=6, value=None)

# Add headers
ws.cell(row=1, column=5, value='General_Sentiment')
ws.cell(row=1, column=6, value='Detailed_Emotion')

# Process each row
for row_idx in range(2, ws.max_row + 1):
    q1 = ws.cell(row=row_idx, column=1).value
    q2 = ws.cell(row=row_idx, column=2).value
    q3 = ws.cell(row=row_idx, column=3).value
    general, detailed = classify_row(q1, q2, q3)
    ws.cell(row=row_idx, column=5, value=general)
    ws.cell(row=row_idx, column=6, value=detailed)

wb.save('data/Written Questions_classified.xlsx')
print("Saved to: Written Questions_classified.xlsx")

# ============ SUMMARY ============
generals = []
details = []
for row_idx in range(2, ws.max_row + 1):
    generals.append(ws.cell(row=row_idx, column=5).value)
    details.append(ws.cell(row=row_idx, column=6).value)

print("=" * 50)
print("=== General Sentiment Distribution ===")
print("=" * 50)
for k, v in Counter(generals).most_common():
    bar = '#' * int(v / len(generals) * 40)
    print(f"  {k:10s}: {v:3d} ({v/len(generals)*100:5.1f}%) {bar}")

print()
print("=" * 50)
print("=== Detailed Emotion Distribution ===")
print("=" * 50)
for k, v in Counter(details).most_common():
    bar = '#' * int(v / len(details) * 40)
    print(f"  {k:12s}: {v:3d} ({v/len(details)*100:5.1f}%) {bar}")

print(f"\nTotal rows processed: {len(generals)}")

# Show cross-tabulation
print()
print("=" * 50)
print("=== Cross-Tab: General x Detailed ===")
print("=" * 50)
cross = {}
for g, d in zip(generals, details):
    cross[(g, d)] = cross.get((g, d), 0) + 1
for (g, d), count in sorted(cross.items(), key=lambda x: -x[1]):
    print(f"  {g:10s} + {d:12s} = {count}")

# Sample rows
print()
print("=== Sample Classifications (first 12 rows) ===")
for row_idx in range(2, min(ws.max_row + 1, 14)):
    q1 = str(ws.cell(row=row_idx, column=1).value or '')[:40]
    q3 = str(ws.cell(row=row_idx, column=3).value or '')[:40]
    gen = ws.cell(row=row_idx, column=5).value
    det = ws.cell(row=row_idx, column=6).value
    print(f"  Row {row_idx:3d}: [{gen:8s} | {det:12s}] Q1: {q1}...")
